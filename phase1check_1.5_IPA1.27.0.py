# -*- coding:utf-8 -*-

from subprocess import PIPE
from selenium import webdriver
from time import sleep
from openpyxl import workbook , load_workbook
import os ,ytFuntion ,re ,subprocess ,time ,shutil ,zipfile
import xml.etree.cElementTree as ET

#加入VIP站點第五六條線,每完成一次就重新下載(用來更新表單已存在序列號資訊)
#1.27.1 加入新資料夾路徑
def strMix(phaseNumber):
   try:
      strStart = AMWSDCStr[AMWSDCStr.find(phase1["{}".format(phaseNumber)])-7:\
                           AMWSDCStr.find(phase1["{}".format(phaseNumber)])]
      if strStart[0] == "t":
         strStart = "h" + strStart
      strEnd = AMWSDCStr[AMWSDCStr.find(phase1["{}".format(phaseNumber)]) + len(phase1["{}".format(phaseNumber)]):\
                         AMWSDCStr.find(phase1["{}".format(phaseNumber)]) + len(phase1["{}".format(phaseNumber)]) + 47]
   
      strMid = AMWSDCStr[AMWSDCStr.find(phase1["{}".format(phaseNumber)]):\
                         AMWSDCStr.find(phase1["{}".format(phaseNumber)]) + len(phase1["{}".format(phaseNumber)])]
      return "{}{}{}".format(strStart ,strMid ,strEnd)
   except:
      return ""

def osRemove(text):
   if os.path.isfile("{}.xlsx".format(text)):    #先確認檔案是否存在
      os.remove("{}.xlsx".format(text))

def osIsFile(text):
   countTime = 0
   while True:
      if os.path.isfile("{}.xlsx".format(text)):
         break
      elif countTime == 5:
         break        
      else:
         sleep(1)
         countTime = countTime + 1      
   if not os.path.isfile("{}.xlsx".format(text)):
      input("確認google有成功登入後,請按ENTER繼續")
   wbDownload = load_workbook("{}.xlsx".format(text),data_only=True) # 打開一個活頁薄
   wbDownload.save(r"{}.xlsx".format(text)) #Excel公式處理
   return load_workbook("{}.xlsx".format(text))    

print("ipa Phase1線路檢查。")

url_Number = input("序列號:").upper().strip()

testdayFile = time.strftime("%y_%m_%d")  
if not os.path.exists(testdayFile):    #先確認資料夾是否存在
   os.makedirs(testdayFile)
   
wb = load_workbook("設定.xlsx")

sheet = wb["帳號"] # 獲取一張表
for i in range(1 ,len(sheet["B"])+1):
   if str(sheet["B" + str(i)].value).strip() == "google":
      googleAccount = str(sheet["C" + str(i)].value).strip() #帳號相關
      googlePassword = str(sheet["D" + str(i)].value).strip()
      
sheet = wb["url"] # 獲取一張表
for i in range(1 ,len(sheet["B"])+1):
   if str(sheet["B" + str(i)].value).strip() == "APKIOS":
      googleApkUrl = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "線路整合":
      googleSpecUrl = str(sheet["D" + str(i)].value).strip()
   if str(sheet["B" + str(i)].value).strip() == "商戶站點地址":
      googleShopUrl = str(sheet["D" + str(i)].value).strip()
      
ogCwd = os.getcwd()
os.chdir(testdayFile)

options = webdriver.ChromeOptions()
prefs = {"profile.default_content_settings.popups": 0, "download.default_directory": os.getcwd()}
options.add_experimental_option("prefs", prefs)

os.chdir(ogCwd)

testWeb = ytFuntion.test_web(webdriver.Chrome(executable_path='chromedriver.exe', chrome_options=options))
testWeb.webDriver.get(googleApkUrl) #目標網址

os.chdir(testdayFile)

osRemove("DEV-ipa%2Fapk 下載位置")
osRemove("線路整合")
osRemove("商戶站點地址")
  
testWeb.elementSendKeys("identifierId" ,1 ,text = googleAccount)#帳號
testWeb.elementClick("div[id='identifierNext'] content[class='CwaK9'] span[class='RveJvd snByac']" ,6)
testWeb.elementSendKeys("input[type=password]" ,6 ,text = googlePassword)#密碼
testWeb.elementClick("div[id='passwordNext'] content[class='CwaK9'] span[class='RveJvd snByac']" ,6)

wbDownload = osIsFile("DEV-ipa%2Fapk 下載位置")
sheetDownload = wbDownload["下載地點"] # 獲取一張表

testWeb.webDriver.get(googleSpecUrl) #目標網址
wbSpec = osIsFile("線路整合")
sheetSpecVIP = wbSpec["VIP線路"] # 獲取一張表
sheetSpec = wbSpec["環境線路"] # 獲取一張表

testWeb.webDriver.get(googleShopUrl) #目標網址
wbShop = osIsFile("商戶站點地址")
sheetShop = wbShop["站點資料"] # 獲取一張表
errorCount = 0

while True:
   error = []
   specDB = ""
   phase1 = {"1":""}
   url = ""
   appName = ""
   specDBList = {"1":"第一套" ,"2":"第二套" ,"3":"第三套" ,"4":"第四套" ,"5":"第五套" ,"6":"第六套","99":"無找到對應套數"}
   urlCheck = ""
   url_name = ""
   if os.path.exists(url_Number):    #先確認資料夾是否存在
      shutil.rmtree(url_Number)
   if url_Number == " ":
      url_Number = input("序列號:").upper().strip()
      osRemove("DEV-ipa%2Fapk 下載位置")
      osRemove("線路整合")
      osRemove("商戶站點地址")

      testWeb.webDriver.get(googleApkUrl) #目標網址
      wbDownload = osIsFile("DEV-ipa%2Fapk 下載位置")
      sheetDownload = wbDownload["下載地點"] # 獲取一張表

      testWeb.webDriver.get(googleSpecUrl) #目標網址
      wbSpec = osIsFile("線路整合")
      sheetSpecVIP = wbSpec["VIP線路"] # 獲取一張表
      sheetSpec = wbSpec["環境線路"] # 獲取一張表

      testWeb.webDriver.get(googleShopUrl) #目標網址
      wbShop = osIsFile("商戶站點地址")
      sheetShop = wbShop["站點資料"] # 獲取一張表
   if url_Number == "Q": #輸入Q則結束
      break

   for i in range(1 ,len(sheetShop["E"])+1):
      if sheetShop["E{}".format(i)].value == None:
         break
      if url_Number == sheetShop["E" + str(i)].value.strip():
         specDB = str(sheetShop["B{}".format(i)].value).strip() #從商戶地址取得套數
         break
   if specDB == "" or specDB == None:
      specDB = "99"
      error.append(specDBList[specDB]) #無套數的顯示
   phase1 = {"1":""}
   VIPcheck = " "
   for i in range(1 ,len(sheetSpecVIP["B"])+1):
      if sheetSpecVIP["B{}".format(i)].value == None:
         break
      elif url_Number == sheetSpecVIP["B{}".format(i)].value: #VIP站點的phase1
         dbNumber = sheetSpecVIP["C{}".format(i)].value
         VIPcheck = "VIP"
         phase1 = {"1":sheetSpecVIP["J{}".format(i)].value ,"2":sheetSpecVIP["K{}".format(i)].value ,\
                   "3":sheetSpecVIP["L{}".format(i)].value ,"4":sheetSpecVIP["M{}".format(i)].value ,\
                   "5":sheetSpecVIP["N{}".format(i)].value ,"6":sheetSpecVIP["O{}".format(i)].value}
         break

   for i in range(1 ,len(sheetSpec["A"])+1):
      if sheetSpec["A{}".format(i)].value == None:
         break
      if specDBList[specDB] == sheetSpec["A{}".format(i)].value:
         ftpUrl = sheetSpec["G{}".format(i)].value #FTP位置
         if phase1["1"] == "": #非VIP站點的phase1
            phase1 = {"1":sheetSpec["H{}".format(i)].value ,"2":sheetSpec["I{}".format(i)].value ,"3":sheetSpec["J{}".format(i)].value ,"4":sheetSpec["K{}".format(i)].value}
            
   if phase1["1"] != "": #第五六條線路為NONE則置換
      try:
         if phase1["5"] == None:
            phase1["5"] = ""
         if phase1["6"] == None:
            phase1["6"] = ""
         '''if phase1["5"] != "" or phase1["6"] != "":
            input("中福VIP出現,請通知YT")'''
      except:
         pass
      
   for i in range(1 ,len(sheetDownload["A"])+1): #取得表單下載路徑的起始位置
      try:
         if sheetDownload["A" + str(i)].value.strip() == "A001":
            startNumber = i
      except:
         pass
       
   for i in range(int(startNumber) ,len(sheetDownload["A"])+1): #序號相關取得
      if sheetDownload["A{}".format(i)].value == None:
         break
      if url_Number == sheetDownload["A{}".format(i)].value.strip():
         url = sheetDownload["D{}".format(i)].value.strip()
         url = url.strip()
         if ftpUrl not in url:
            urlCheck = "{} 下載位置錯誤".format(url)
            error.append(urlCheck)
            break
         elif url == None or url == "":
            error.append("下載連結空白。")
            break
         url_name = sheetDownload["B{}".format(i)].value
         appName = sheetDownload["C{}".format(i)].value
         urlCheck = "{} 下載位置正確".format(url)
      
         countTime = 0
         if os.path.isfile(appName + ".ipa"):    #先確認檔案是否存在
            os.remove(appName + ".ipa")
         testWeb.webDriver.get(re.sub("plist" ,"ipa" ,url))
         while True:
            if os.path.isfile(appName +".ipa"):
               break
            elif countTime == 10:
               break        
            else:
               sleep(1)
               countTime = countTime + 1 
   #sleep(5)
   countTime = 0
   if len(error) > 0 or url == None or url == "":
      deZip = False
   else:
      deZip = True
      
   while deZip:
      item = "{}\\{}.ipa".format(os.getcwd() ,appName)
      try:
         zpf = zipfile.ZipFile(item) # zip object
      except:
         print("An error occurred in importing the file as a ZIP Archive!")

      try:
         AMWSDCNameSpace = "Payload/AMWSDC.app/AMWSDC"
         AMWSDC = zpf.open(AMWSDCNameSpace) # opens binary AMWSDC
         AMWSDC_r = AMWSDC.read()
         AMWSDCStr = AMWSDC_r.decode(encoding="utf-8" ,errors='ignore')
         break
      except:
         pass
         
      try:
         AMWSDCNameSpace = "Payload/AFNetworking.app/AMWSDC"
         AMWSDC = zpf.open(AMWSDCNameSpace) # opens binary AMWSDC
         AMWSDC_r = AMWSDC.read()
         AMWSDCStr = AMWSDC_r.decode(encoding="utf-8" ,errors='ignore')
         break
      except:
         print("找不到AMWSDC!")
         
      sleep(1)
      countTime = countTime + 1
      if countTime == 5:
         break
         error.append("無此檔案。")
         
   requestUrl = {"requestUrl":strMix(1) ,"requestUrl2":strMix(2) ,"requestUrl3":strMix(3) ,\
                 "requestUrl4":strMix(4)  ,"requestUrl5":strMix(5)  ,"requestUrl6":strMix(6)}
   

   requestUrl1 ,requestUrl2 ,requestUrl3 ,requestUrl4 ,requestUrl5 ,requestUrl6 = "NG" ,"NG" ,"NG" ,"NG" ,"NG" ,"NG" 
   for i in phase1:
      if requestUrl["requestUrl5"] != "" and requestUrl["requestUrl6"] != "": #判斷ipa中的phase1是否正確
         if phase1[i] in requestUrl["requestUrl"] and url_Number in requestUrl["requestUrl"]:
            requestUrl1 = "OK"
         elif phase1[i] in requestUrl["requestUrl2"] and url_Number in requestUrl["requestUrl2"]:
            requestUrl2 = "OK"
         elif phase1[i] in requestUrl["requestUrl3"] and url_Number in requestUrl["requestUrl3"]:
            requestUrl3 = "OK"
         elif phase1[i] in requestUrl["requestUrl4"] and url_Number in requestUrl["requestUrl4"]:
            requestUrl4 = "OK"
         elif phase1[i] in requestUrl["requestUrl5"] and url_Number in requestUrl["requestUrl5"]:
            requestUrl5 = "OK"
         elif phase1[i] in requestUrl["requestUrl6"] and url_Number in requestUrl["requestUrl6"]:
            requestUrl6 = "OK"
      elif requestUrl["requestUrl5"] != "" and requestUrl["requestUrl6"] == "":
            error.append("有phase5卻沒phase6")
            break
      elif requestUrl["requestUrl5"] == "" and requestUrl["requestUrl6"] != "":
            error.append("有phase6卻沒phase5")
            break
      else:
         if phase1[i] in requestUrl["requestUrl"] and url_Number in requestUrl["requestUrl"]:
            requestUrl1 = "OK"
         elif phase1[i] in requestUrl["requestUrl2"] and url_Number in requestUrl["requestUrl2"]:
            requestUrl2 = "OK"
         elif phase1[i] in requestUrl["requestUrl3"] and url_Number in requestUrl["requestUrl3"]:
            requestUrl3 = "OK"
         elif phase1[i] in requestUrl["requestUrl4"] and url_Number in requestUrl["requestUrl4"]:
            requestUrl4 = "OK"
            
   if url != "":
      print(url_Number) #輸出結果
      print(url_name)
      print("DB:{}".format(specDBList[specDB]))
      print(urlCheck)
      if VIPcheck == "VIP":
         print("{}為VIP站點。".format(url_Number))
      if requestUrl1 != "OK":
         print("{} 錯誤。".format(requestUrl["requestUrl"]))
      else:
         print("{} 線路正確。".format(requestUrl["requestUrl"]))
      if requestUrl2 != "OK":
         print("{} 錯誤。".format(requestUrl["requestUrl2"]))
      else:
         print("{} 線路正確。".format(requestUrl["requestUrl2"]))
      if requestUrl3 != "OK":

         print("{} 錯誤。".format(requestUrl["requestUrl3"]))
      else:
         print("{} 線路正確。".format(requestUrl["requestUrl3"]))
      if requestUrl4 != "OK":
         print("{} 錯誤。".format(requestUrl["requestUrl4"]))
      else:
         print("{} 線路正確。".format(requestUrl["requestUrl4"]))
      if requestUrl["requestUrl6"] != "":
         if requestUrl5 != "OK":
            print("{} 錯誤。".format(requestUrl["requestUrl5"]))
         else:
            print("{} 線路正確。".format(requestUrl["requestUrl5"]))
         if requestUrl6 != "OK":
            print("{} 錯誤。".format(requestUrl["requestUrl6"]))
         else:
            print("{} 線路正確。".format(requestUrl["requestUrl6"]))
            print("{}為中福VIP。".format(url_Number))
      print("FTP: {}".format(ftpUrl))
      url_Number = " "
      errorCount = 0
   else:
      print("無此序列號,重新下載。")
      osRemove("DEV-ipa%2Fapk 下載位置")
      osRemove("線路整合")
      osRemove("商戶站點地址")

      testWeb.webDriver.get(googleApkUrl) #目標網址
      wbDownload = osIsFile("DEV-ipa%2Fapk 下載位置")
      sheetDownload = wbDownload["下載地點"] # 獲取一張表

      testWeb.webDriver.get(googleSpecUrl) #目標網址
      wbSpec = osIsFile("線路整合")
      sheetSpecVIP = wbSpec["VIP線路"] # 獲取一張表
      sheetSpec = wbSpec["環境線路"] # 獲取一張表

      testWeb.webDriver.get(googleShopUrl) #目標網址
      wbShop = osIsFile("商戶站點地址")
      sheetShop = wbShop["站點資料"] # 獲取一張表
      errorCount += 1
      if errorCount == 2:
         print("站點獲取重試失敗。")
         print()
         url_Number = " "
         
   if errorCount == 0:
      if len(error) == 0:
         print()
      else:
         print(error)
         print()
   #break
testWeb.webDriver.close()
